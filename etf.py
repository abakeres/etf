import pandas as pd
import yfinance as yf
import matplotlib.pyplot as plt
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

def get_stock_info(tickers):
    stock_data = {}
    for ticker in tickers:
        t = yf.Ticker(ticker)
        info = t.info
        stock_data[ticker] = info
        print(f"{ticker} loaded - {info.get('shortName', 'N/A')}")
    return stock_data

def get_key_metrics(stock_data):
    metrics = {
        "Current Price": "currentPrice",
        "Market Cap": "marketCap",
        "52W High": "fiftyTwoWeekHigh",
        "52W Low": "fiftyTwoWeekLow",
        "Volume": "volume",
        "Avg. Volume": "averageVolume",
        "Beta": "beta",
        "Dividend Yield": "dividendYield",
        "Ex-Dividend Date": "exDividendDate",
        "Shares Outstanding": "sharesOutstanding",
        "Float": "float",
        "Short % of Float": "shortPercentOfFloat",
    }

    rows = []
    for label, key in metrics.items():
        row = {"Metric": label}
        for ticker in stock_data:
            row[ticker] = stock_data[ticker].get(key, "N/A")
        rows.append(row)
    
    return pd.DataFrame(rows)

def get_valuation(stock_data):
    metrics = {
        "P/E Ratio (TTM)": "trailingPE",
        "Forward P/E": "forwardPE",
        "PEG Ratio": "pegRatio",
        "Price-to-Sales": "priceToSalesTrailing12Months",
        "Price-to-Book": "priceToBook",
        "EV/EBITDA": "enterpriseToEbitda",
        "EV/Revenue": "enterpriseToRevenue",
        "Enterprise Value": "enterpriseValue",
    }

    rows = []
    for label, key in metrics.items():
        row = {"Metric": label}
        for ticker in stock_data:
            row[ticker] = stock_data[ticker].get(key, "N/A")
        rows.append(row)

    return pd.DataFrame(rows)

def get_financials(stock_data):
    metrics = {
        "Revenue (TTM)": "totalRevenue",
        "Revenue Growth (YoY)": "revenueGrowth",
        "Gross Margin": "grossMargins",
        "Profit Margin": "profitMargins",
        "EPS (TTM)": "trailingEps",
        "EPS Growth": "earningsGrowth",
        "Debt-to-Equity": "debtToEquity",
        "Free Cash Flow": "freeCashflow",
        "Return on Equity": "returnOnEquity",
        "Return on Assets": "returnOnAssets",
    }

    rows = []
    for label, key in metrics.items():
        row = {"Metric": label}
        for ticker in stock_data:
            row[ticker] = stock_data[ticker].get(key, "N/A")
        rows.append(row)
    
    return pd.DataFrame(rows)

def get_analyst_ratings(tickers):
    rows = []
    for ticker in tickers:
        t = yf.Ticker(ticker)
        rec = t.recommendations_summary
        if rec is not None and not rec.empty:
            latest = rec.iloc[0]
            rows.append({
                "Ticker": ticker,
                "Strong Buy": latest.get("strongBuy", 0),
                "Buy": latest.get("buy", 0),
                "Hold": latest.get("hold", 0),
                "Underperform": latest.get("underperform", 0),
                "Sell": latest.get("sell", 0),
             })
        else:
            rows.append({
                "Ticker": ticker,
                "Strong Buy": "N/A",
                "Buy": "N/A",
                "Hold": "N/A",
                "Underperform": "N/A",
                "Sell": "N/A",
            })
    df = pd.DataFrame(rows)
    if "Strong Buy" in df.columns:
        df["Total"] = df[["Strong Buy", "Buy", "Hold", "Underperform", "Sell"]].apply(pd.to_numeric, errors="coerce").sum(axis=1)
    return df

def get_news(tickers):
    import anthropic
    from newspaper import Article

    client = anthropic.Anthropic()
    rows = []

    for ticker in tickers:
        t = yf.Ticker(ticker)
        news = t.news[:5]

        for article in news:
            url = article.get("link", "")
            headline = article.get("title", "N/A")
            source = article.get("publisher", "N/A")
            published = pd.to_datetime(article.get("providerPublishTime", 0), unit="s").strftime("%Y-%m-%d")

            try:
                a = Article(url)
                a.download()
                a.parse()
                article_text = a.text[:3000]

                response = client.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=300,
                    messages=[{"role": "user", "content": f"Summarize this article in a short paragraph: {article_text}"}]
                )
                summary = response.content[0].text
            except:
                summary = "Could not fetch article"

            rows.append({
                "Ticker": ticker,
                "Headline": headline,
                "Source": source,
                "Date": published,
                "URL": url,
                "Summary": summary
            })
    return pd.DataFrame(rows)

def get_stock_performance(tickers):
    performance_data = {}
    for ticker in tickers:
        t = yf.Ticker(ticker)
        hist = t.history(period="1y")
        hist = hist[["Close"]]
        hist.index = hist.index.tz_localize(None)
        hist["Normalized"] = (hist["Close"] / hist["Close"].iloc[0]) * 100
        performance_data[ticker] = hist
        total_return = ((hist["Close"].iloc[-1] - hist["Close"].iloc[0]) / hist["Close"].iloc[0]) * 100
        print(f"    {ticker}: {total_return:.1f}% return over past year")
    return performance_data

def export_stock_excel(tickers, stock_data, key_metrics_df, valuation_df, financials_df, analyst_df, news_df, performance_df):
    print("\nExporting to Excel...")

    with pd.ExcelWriter("stock_comparison.xlsx", engine="openpyxl") as writer:

        key_metrics_df.to_excel(writer, sheet_name="Key Metrics", index=False)

        valuation_df.to_excel(writer, sheet_name="Valuation", index=False)

        financials_df.to_excel(writer, sheet_name="Financials", index=False)

        analyst_df.to_excel(writer, sheet_name="Analyst Ratings", index=False)

        news_df.to_excel(writer, sheet_name="News", index=False)

        workbook = writer.book
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        center = Alignment(horizontal="center")

        for sheet_name in ["Key Metrics", "Valuation", "Financials", "Analyst Ratings", "News"]:
            ws = workbook[sheet_name]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4
            
        ws_news = workbook["News"]
        url_col = None
        for cell in ws_news[1]:
            if cell.value =="URL":
                url_col = cell.column
                break
        if url_col:
            for row in range(2, ws_news.max_row +1):
                cell = ws_news.cell(row=row, column=url_col)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")

        fig, ax = plt.subplots(figsize=(10, 6))
        rating_cols = ["Strong Buy", "Buy", "Hold", "Underperform", "Sell"]
        analyst_chart_df = analyst_df.set_index("Ticker")[rating_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
        analyst_chart_df.plot(kind="bar", stacked=True, ax=ax, color=["#2ecc71", "#27ae60", "#f39c12", "#e74c3c", "#c0392b"])
        ax.set_title("Analyst Ratings Breakdown")
        ax.set_xlabel("Stock")
        ax.set_ylabel("Number of Analysts")
        ax.legend(title="Rating")
        plt.xticks(rotation=0)
        plt.tight_layout()
        img_bytes = io.BytesIO()
        plt.savefig(img_bytes, format="png")
        img_bytes.seek(0)
        plt.close()

        ws_analyst = workbook["Analyst Ratings"]
        img = Image(img_bytes)
        ws_analyst.add_image(img, f"A{len(analyst_df) + 4}")

        ws_charts = workbook.create_sheet("Charts")
        fig2, ax2 = plt.subplots(figsize=(12, 6))
        for ticker in tickers:
            performance_data[ticker]["Normalized"].plot(ax=ax2, label=ticker)
        ax2.set_title("1 Year Price Performance (Nomalized to 100)")
        ax2.set_xlabel("Date")
        ax2.set_ylabel("Value (Starting at 100)")
        ax2.legend(title="Stock")
        plt.tight_layout()
        img_bytes2 = io.BytesIO()
        plt.savefig(img_bytes2, format="png")
        img_bytes2.seek(0)
        plt.close()
        img2 = Image(img_bytes2)
        ws_charts.add_image(img2, "A1")
    
    print("Done! File saved as stock_comparison.xlsx")

mode = input("Do you want to compare ETFs or Stocks? (etf/stock): ").strip().lower()

if mode not in ["etf", "stock"]:
    print("Invalid choice. Please enter 'etf' or 'stock'. ")
    exit()

user_input = input(f"Enter {'ETF' if mode == 'etf' else 'stock'} tickers seperated by commas: ")
tickers = [t.strip().upper() for t in user_input.split(",")]

if mode == "etf":
    all_holdings = {}

    for etf in tickers:
        ticker = yf.Ticker(etf)
        holdings = ticker.funds_data.top_holdings
        holdings = holdings.reset_index()
        holdings.columns = ["Ticker", "Name", "Weight"]
        all_holdings[etf] = holdings
        print(f"{etf} loaded - {len(holdings)} holdings found")
        for _, row in holdings.iterrows():
            print(f"    {row['Ticker']} - {row['Name']}: {row['Weight']:.1%}")

    print("\nALL ETFS loaded successfully!")

    print("\nOverlap Analysis:")

    etf_list = list(all_holdings.keys())

    for i in range(len(etf_list)):
        for j in range(i+1, len(etf_list)):
            etf1 = etf_list[i]
            etf2 = etf_list[j]

            tickers1 = set(all_holdings[etf1]["Ticker"])
            tickers2 = set(all_holdings[etf2]["Ticker"])

            overlap = tickers1.intersection(tickers2)
            overlap_count = len(overlap)

            weight1 = all_holdings[etf1][all_holdings[etf1]["Ticker"].isin(overlap)]["Weight"].sum()
            weight2 = all_holdings[etf2][all_holdings[etf2]["Ticker"].isin(overlap)]["Weight"].sum()

            print(f"{etf1} vs {etf2}: {overlap_count} stocks in common")
            print(f"    Overlap weight in {etf1}: {weight1:.1%}")
            print(f"    Overlap weigth in {etf2}: {weight2:.1%}")

    print("\nSector Allocation:")

    sector_data = {}

    for etf in tickers:
        ticker = yf.Ticker(etf)
        sectors = ticker.funds_data.sector_weightings
    
        cleaned = {k.replace("_", " ").title(): round(v * 100, 2) for k, v in sectors.items()}
        sector_data[etf] = cleaned

        print(f"\n{etf} Sector Breakdown:")
        for sector, weight in sorted(cleaned.items(), key=lambda x: x[1], reverse=True):
            print(f"    {sector}: {weight}%")

    print("\nPulling performance data...")

    performance_data = {}

    for etf in tickers:
        ticker = yf.Ticker(etf)
        hist = ticker.history(period="1y")
        hist = hist[["Close"]]
        hist.index = hist.index.tz_localize(None)

        hist["Normalized"] = (hist["Close"] / hist["Close"].iloc[0]) * 100
        performance_data[etf] = hist
        total_return = ((hist["Close"].iloc[-1] - hist["Close"].iloc[0]) / hist["Close"].iloc[0]) * 100
        print(f"    {etf}: {total_return:.1f}% return over past year")

    print("\nExporting to Excel...")

    with pd.ExcelWriter("etf_comparison.xlsx", engine="openpyxl") as writer:

        holdings_combined = pd.DataFrame()
        for etf in tickers:
            df = all_holdings[etf].copy()
            df["ETF"] = etf
            holdings_combined = pd.concat([holdings_combined, df])
        holdings_combined.to_excel(writer, sheet_name="Holdings", index=False)

        overlap_rows = []
        for i in range(len(etf_list)):
            for j in range(i+1, len(etf_list)):
                etf1 = etf_list[i]
                etf2 = etf_list[j]
                tickers1 = set(all_holdings[etf1]["Ticker"])
                tickers2 = set(all_holdings[etf2]["Ticker"])
                overlap = tickers1.intersection(tickers2)
                weight1 = all_holdings[etf1][all_holdings[etf1]["Ticker"].isin(overlap)]["Weight"].sum()
                weight2 = all_holdings[etf2][all_holdings[etf2]["Ticker"].isin(overlap)]["Weight"].sum()
                overlap_rows.append({
                    "ETF Pair": f"{etf1} vs {etf2}",
                    "Stocks in Common": len(overlap),
                    f"Overlap Weight in {etf1}": f"{weight1:.1%}",
                    f"Overlap Weight in {etf2}": f"{weight2:.1%}",
                    "Shared Tickers": ", ".join(sorted(overlap))
                })
        overlap_df = pd.DataFrame(overlap_rows)
        overlap_df.to_excel(writer, sheet_name="Overlap", index=False)

        sector_df = pd.DataFrame(sector_data).fillna(0)
        sector_df.index.name = "Sector"
        sector_df.to_excel(writer, sheet_name="Sectors")

        workbook = writer.book
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        center = Alignment(horizontal="center")

        ws_summary = workbook.create_sheet("Summary", 0)

        ws_summary["A1"] = "ETF Comparison Summary"
        ws_summary["A1"].font = Font(bold=True, size=16, color="1F4E79")
        ws_summary["A2"] = f"ETFS Analyzed: {', '.join(tickers)}"
        ws_summary["A2"].font = Font(italic=True, size=11)
        ws_summary["A3"] = ""

        row = 4

        ws_summary.cell(row=row, column=1, value="Top Holding in ETF").font = Font(bold=True, size=12)
        row += 1
        for etf in tickers:
            top = all_holdings[etf].iloc[0]
            ws_summary.cell(row=row, column=1, value=f"{etf}: {top['Name']} ({top['Weight']:.1%})")
            row += 1

        row += 1

        ws_summary.cell(row=row, column=1, value="Larget Sector per ETF").font = Font(bold=True, size=12)
        row += 1
        for etf in tickers:
            top_sector = max(sector_data[etf], key=sector_data[etf].get)
            top_weight = sector_data[etf][top_sector]
            ws_summary.cell(row=row, column=1, value=f"{etf}: {top_sector} ({top_weight}%)")
            row += 1

        row += 1

        ws_summary.cell(row=row, column=1, value="Overlap Analysis").font = Font(bold=True, size=12)
        row += 1
        for i in range(len(etf_list)):
            for j in range(i + 1, len(etf_list)):
                etf1 = etf_list[i]
                etf2 = etf_list[j]
                tickers1 = set(all_holdings[etf1]["Ticker"])
                tickers2 = set(all_holdings[etf2]["Ticker"])
                overlap = tickers1.intersection(tickers2)
                weight1 = all_holdings[etf1][all_holdings[etf1]["Ticker"].isin(overlap)]["Weight"].sum()
                weight2 = all_holdings[etf2][all_holdings[etf2]["Ticker"].isin(overlap)]["Weight"].sum()
                ws_summary.cell(row=row, column=1, value=f"{etf1} vs {etf2}: {len(overlap)} stocks in common - {weight1:.1%} of {etf1}, {weight2:.1%} of {etf2}")
                row += 1

        ws_summary.column_dimensions["A"].width = 70

        fig, ax = plt.subplots(figsize=(12, 6))
        sector_df_chart = pd.DataFrame(sector_data). fillna(0)
        sector_df_chart.plot(kind="bar", ax=ax)
        ax.set_title("Sector Allocation by ETF")
        ax.set_xlabel("Sector")
        ax.set_ylabel("Weight (%)")
        ax.legend(title="ETF")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()

        img_bytes = io.BytesIO()
        plt.savefig(img_bytes, format="png")
        img_bytes.seek(0)
        plt.close()

        ws_chart = workbook.create_sheet("Charts")
        img = Image(img_bytes)
        ws_chart.add_image(img, "A1")

        fig2, ax2 = plt.subplots(figsize=(12, 6))
        for etf in tickers:
            performance_data[etf]["Normalized"].plot(ax=ax2, label=etf)
        ax2.set_title("1 Year Price Performance (Normalized to 100)")
        ax2.set_xlabel("Date")
        ax2.set_ylabel("Value (Starting at 100)")
        ax2.legend(title="ETF")
        plt.tight_layout()

        img_bytes2 = io.BytesIO()
        plt.savefig(img_bytes2, format="png")
        img_bytes2.seek(0)
        plt.close()

        img2 = Image(img_bytes2)
        ws_chart.add_image(img2, "A22")

        for sheet_name in ["Holdings", "Overlap", "Sectors"]:
            ws = workbook[sheet_name]

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    print("Done! File saved as etf_comparison.xlsx")

if mode == "stock":
    print(f"\nLoading data for: {', '.join(tickers)}")

    stock_data = get_stock_info(tickers)

    print(("\nBuilding Key Metrics..."))
    key_metrics_df = get_key_metrics(stock_data)

    print("Building Valuation...")
    valuation_df = get_valuation(stock_data)
    print("Building Financials...")
    financials_df = get_financials(stock_data)

    print("Fetching Analyst Ratings...")
    analyst_df = get_analyst_ratings(tickers)

    print("Fetching News and Summarizing...")

    news_df = get_news(tickers)
    print("Fetching Performance Data...")
    performance_data = get_stock_performance(tickers)

    export_stock_excel(tickers, stock_data, key_metrics_df, valuation_df, financials_df, analyst_df, news_df, performance_data)